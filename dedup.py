####!/usr/bin/python3

import tkinter as tk
from tkinter import font as tkFont
from tkinter.filedialog import askopenfilename
import openpyxl
import pickle
import sys
import re

essVal = [[[],[],[]]]
nnvar = 0

class var:
  def __init__(self):
    self.fname = "x"
    self.skuID = []
    self.skuTitle = []
    self.ess = []
    self.labelDetail = []
    self.nn = 0
    self.n1 = 0
    self.n2 = 0
    self.n3 = 0



class mainFunc:
  def __init__(self):
    pass

  def loadfile(self):
    tk.Tk().withdraw()
    filename = askopenfilename()
    print(filename)
    if filename == "":
      sys.exit()
    #skufile = "C:\\Storage\\Nexrise\\work_2020-07-01\\N92_Musab_Nexrise_Training_DataSet4.xlsx"
    self.skufile = filename
    self.excelload()
    self.storedata()

  def excelload(self):
    main1 = 2
    main2 = 3
    main3 = 9
    col1 = 4
    col2 = 5
    col3 = 6
    col4 = 7
    col5 = 8
    col6 = 10


    self.skuID = []
    self.skuTitle = []
    self.ess = []
    wb = openpyxl.load_workbook(self.skufile)
    sheet = wb["Sheet1"]
    print(str(sheet))
    for row in range(2, sheet.max_row + 1):  # sheet.max_row + 1
        data = sheet.cell(row=row, column=main1)
        cell = data.value
        self.skuID.append(cell)

    for row in range(2, sheet.max_row + 1):  # sheet.max_row + 1
        data = sheet.cell(row=row, column=main2)
        cell = data.value
        self.skuTitle.append(cell)

    for row in range(2, sheet.max_row + 1):  # sheet.max_row + 1
        data = sheet.cell(row=row, column=main3)
        cell = data.value
        self.ess.append(cell)
        
    print("Found " + str(len(self.skuID)) + " raw data")
    print("\nskuID",self.skuID[:3],"\nskuTitle",self.skuTitle[:3],"\ness",self.ess[:3])

    self.labelDetail = []
    for row in range(2, sheet.max_row + 1):  # sheet.max_row + 1
        labels = str(sheet.cell(row=row, column=col1).value) + " | " + str(sheet.cell(row=row, column=col2).value) + " | " + str(sheet.cell(row=row, column=col3).value) + " | " + str(sheet.cell(row=row, column=col4).value) + " | " + str(sheet.cell(row=row, column=col5).value) + " | " + str(sheet.cell(row=row, column=col6).value)
        self.labelDetail.append(labels)
    len(self.labelDetail)
    print("labelDetail",self.labelDetail[:3])



  def storedata(self):
    print("storing pickle data")
    dbfile = open('data', 'wb')
    db = [[self.skufile],[self.skuID],[self.skuTitle],[self.ess],[self.labelDetail]]
    pickle.dump(db, dbfile)
    dbfile.close()

    print("storing pickle empty labels")
    var.results = []
    for ln in range(0, len(self.skuTitle)):
          var.results.append([])
    dbfile = open('labelled', 'wb')
    db = var.results
    pickle.dump(db, dbfile)
    dbfile.close()
    print(db[:10])

    print("reading stored labels")
    dbfile = open('labelled', 'rb')
    print(pickle.load(dbfile)[:10])
    dbfile.close()





  def loaddata(self):
    try:
      print("reading pickle data")
      dbfile = open('data', 'rb')
      db = pickle.load(dbfile)
      dbfile.close()
      var.fname = db[0][0].split("/")[-1]
      print(len(db),var.fname)
      var.skuID = db[1][0]
      var.skuTitle = db[2][0]
      var.ess = db[3][0]
      var.labelDetail = db[4][0]
      var.nn = 0
      if var.nn == 0:
          var.n1 = ""
      else:
          var.n1 = var.skuTitle[var.nn-1]
          print(var.n1)
      var.n2 = var.skuTitle[var.nn]
      print(var.n2)
      var.n3 = var.skuTitle[var.nn+1]
      print(var.n3)
      var.skulength = len(var.skuTitle)
      ctrlFunc().cleanEss()

      print("reading pickle labels")
      dbfile = open('labelled', 'rb')
      var.results = pickle.load(dbfile)
      dbfile.close()
      print(var.results[:10])

    except Exception as e:
      print("loading pickle failed", e)
      var.fname = ""
      var.skuID = [[],[],[]]
      var.skuTitle = [[],[],[]]
      var.ess = [[],[],[]]
      var.labelDetail = ["","",""]
      var.nn = 0
      var.n1 = ""
      var.n2 = ""
      var.n3 = ""
      var.essVal = [[[],[],[]]]
      var.skulength = ""
      var.skudetail = ""

  def findDup(self, s0, s1, s2):
    ss0 = set(s0.lower().split(" "))
    ss1 = set(s1.lower().split(" "))
    ss2 = set(s2.lower().split(" "))
    common1 = ss0 & ss1
    common2 = ss1 & ss2
    cm1 = [[],[]]
    for c1 in common1:
      res = s0.lower().find(c1)
      if res > -1:
        cm1[0].append(res)
        cm1[1].append(c1)
    cm2 = [[],[]]
    for c2 in common2:
      res = s2.lower().find(c2)
      if res > -1:
        cm2[0].append(res)
        cm2[1].append(c2)
    return cm1, cm2

  def txtdetail(self):
    var.skudetail = "[" + str(var.nn) + "/" + str(var.skulength) + "] "
    var.skudetail += var.n2 + "\n\n"
    for ev in var.essVal[var.nn]:
      if len(ev) > 0:
        try:
          evn = re.split(r'\s\:\:\:\:\s|\:',ev)[1]
          if evn.find('", "') > -1:
              evn = evn.split('", "')[0]
          if evn.find('""') > -1:
              evn = evn.split('""')[0]
          var.skudetail += " " + re.sub(r'[^A-Za-z0-9\.]', ' ', evn).lower() + " \n"
        except:
          print("Failed splitting text on txtdetail function")
    var.skudetail += "\n" + var.labelDetail[var.nn] + "\n\n"

    ## Add next previous skus
    var.skudetail += var.n1 + "\n" + var.n2 + "\n" + var.n3

    ## Add labelled
    var.skudetail += "\n\n" + str(var.results[var.nn])
        
        
  def changetxtdetail2(self):
    var.T.delete('1.0', tk.END)
    self.txtdetail()
    var.T.insert(tk.END, var.skudetail)
    
    ## highlight matches
    var.spldetail = var.skudetail.split("\n\n")[1].split("\n")
    #n2s = n2.split(" ")
    n2s = re.split(r'[^A-Za-z0-9\.]', var.n2)
    for ns in n2s:
      nns = 3
      for spld in var.spldetail:
        ns = ns.lower()
        #print(ns,spld)
        if spld.find(ns) > -1:
          posnsx = str(nns) + "." + str(spld.find(ns))
          posnsy = str(nns) + "." + str(spld.find(ns) + len(ns))
          #print(ns, posnsx, posnsy)
          var.T.tag_add("start", posnsx, posnsy)
          var.T.tag_config("start", background="black",foreground="yellow")
        nns += 1
    dup1, dup2 = self.findDup(var.n1,var.n2,var.n3)
    n = 0
    for d1 in dup1[0]:
      posnsx = "9." + str(d1)
      posnsy = "9." + str(d1 + len(dup1[1][n]))
      n += 1
      var.T.tag_add("start", posnsx, posnsy)
      var.T.tag_config("start", background="black",foreground="yellow")
    n = 0
    for d2 in dup2[0]:
      posnsx = str("11.") + str(d2)
      posnsy = str("11.") + str(d2 + len(dup2[1][n]))
      n += 1
      var.T.tag_add("start", posnsx, posnsy)
      var.T.tag_config("start", background="black",foreground="yellow")

      









class ctrlFunc():
  def cleanEss(self):
    print("splitting ess values")
    var.essVal = []
    for es in var.ess:
      try:
        #es = es.split(")(")
        #es = es.split("////")
        #es = re.split(r'\)\(|\/\/\/\/',es)
        es = re.split(r'\}\-\-\{',es)
        if len(es) == 1:
          #print(es)
          es = [es[0],"###!","###!"]
          #print(es)
        var.essVal.append(es)
      except:
        print("Ess split failed")
        var.essVal.append([])
    print("var.essVal len in", len(var.essVal))
  










class tkApp:
  def __init__(self, root):
    self.root = root

    self.f1()
    self.f2()
    self.f3()
    mainFunc().changetxtdetail2()


  def jumpnn(self):
    varnn = int(self.name_entry.get())
    var.nn = int(varnn) - 1
    self.next()

  def f1(self):
    self.topframe = tk.Frame(self.root, width = 1200, height = 80, background="green")
    self.topframe.pack(side="top")
    self.topframe.pack_propagate(0)
    self.topbtn()

  def topbtn(self):
    name_label = tk.Label(self.topframe, text = 'Line#', background="green")
    name_label.pack(side="left")
    self.name_entry = tk.Entry(self.topframe, width=8)
    self.name_entry.pack(side="left")

    btn4 = tk.Button(self.topframe, text="Go", fg="black", activebackground = "white", command=self.jumpnn, height = 1)
    btn4.pack(side="left", padx = (0,25))
    
    btn4 = tk.Button(self.topframe, text="Load File", fg="black", activebackground = "white", command=self.loadfile, height = 1)
    btn4.pack(side="left")
    print(var.fname)
    name_label = tk.Label(self.topframe, text = str('current loaded file : '+ var.fname))
    name_label.pack(side="left")

    btn4 = tk.Button(self.topframe, text="Reload File", fg="black", activebackground = "white", command=self.reload, height = 1)
    btn4.pack(side="left" , padx = 25)



  def f2(self):
    self.frame1 = tk.Frame(self.root, width = 1310, height = 190, background="blue")
    self.frame1.pack(side="top")
    self.frame1.pack_propagate(0)
    self.frames2a()
    self.frames2b()

  def frames2a(self):
    self.leftframe = tk.Frame(self.frame1,  width = 650, height = 180,background="yellow")
    self.leftframe.pack(side="left")
    self.leftframe.pack_propagate(0)

    btnwidth = 120
    btn4 = tk.Button(self.leftframe, text=var.n1, fg="orange", activebackground = "orange", command=self.previous, height = 2, width = btnwidth, wraplength=500)
    btn4.pack()
    btn4 = tk.Button(self.leftframe, text=var.n2, fg="blue", activebackground = "blue", height = 2, width = btnwidth, wraplength=500)
    btn4.pack()
    btn4 = tk.Button(self.leftframe, text=var.n3, fg="orange", activebackground = "orange", command=self.next, height = 2, width = btnwidth, wraplength=500)
    btn4.pack()

  def frames2b(self):
    self.rightframe = tk.Frame(self.frame1,  width = 650, height = 180,background="orange")
    self.rightframe.pack(side="right")
    self.rightframe.pack_propagate(0)
    btnwidth = 120
    btn3 = tk.Button(self.rightframe, text=var.essVal[var.nn][0], fg="black", activebackground = "white", command=self.opt1, height = 2, width = btnwidth, wraplength=580)
    btn3.pack()
    btn3 = tk.Button(self.rightframe, text=var.essVal[var.nn][1], fg="black", activebackground = "white", command=self.opt2, height = 2, width = btnwidth, wraplength=580)
    btn3.pack()
    btn3 = tk.Button(self.rightframe, text=var.essVal[var.nn][2], fg="black", activebackground = "white", command=self.opt3, height = 2, width = btnwidth, wraplength=580)
    btn3.pack()

  def f3(self):
    self.optionframe = tk.Frame(self.root, width = 800, height = 80, background="green")
    self.optionframe.pack()
    self.optionframe.pack_propagate(0)

    self.textframe = tk.Frame(self.root, width = 1200, height = 380, background="brown")
    self.textframe.pack(side="bottom")
    self.textframe.pack_propagate(0)

    self.optionbtn()
    self.txtbox()

  def optionbtn(self):
    btnwidth = 6
    btnwidth2 = 12
    fontsize = 25
    helv36 = tkFont.Font(family='Helvetica', size=10)
    btn4 = tk.Button(self.optionframe, font=helv36, text="<", fg="black", activebackground = "white", command=self.goprev, width = btnwidth,height = 3, wraplength=40)
    btn4.pack(side="left", padx=5)
    btn4 = tk.Button(self.optionframe, font=helv36, text=">", fg="black", activebackground = "white", command=self.gonext, width = btnwidth,height = 3, wraplength=40)
    btn4.pack(side="left", padx=5)
    btn4 = tk.Button(self.optionframe, font=helv36, text="Save List", fg="black", activebackground = "white", command=self.save, width = btnwidth,height = 3, wraplength=40)
    btn4.pack(side="left", padx=5)
    btn4 = tk.Button(self.optionframe, font=helv36, text="Last Work", fg="black", activebackground = "white", command=self.lastwork, width = btnwidth,height = 3, wraplength=40)
    btn4.pack(side="left", padx=5)
    btn4 = tk.Button(self.optionframe, font=helv36, text="Blank", fg="black", activebackground = "white", command=self.blanklabel, width = btnwidth2,height = 3, wraplength=80)
    btn4.pack(side="left", padx=5)
    btn4 = tk.Button(self.optionframe, font=helv36, text="Duplicate as previous", fg="black", activebackground = "white", command=self.dupPrev, width = btnwidth2,height = 3, wraplength=80)
    btn4.pack(side="left", padx=5)
    btn4 = tk.Button(self.optionframe, font=helv36, text="New - True", fg="black", activebackground = "white", command=self.newTrue, width = btnwidth2,height = 3, wraplength=80)
    btn4.pack(side="left", padx=5)

  def txtbox(self):
    var.T = tk.Text(self.textframe, font=(20))
    var.T.pack(fill='x')






  def loadfile(self):
    mainFunc().loadfile()
    var.nn = 1
    self.previous()

  def previous(self):
    self.leftframe.destroy()
    self.rightframe.destroy()
    #print(self.nn, len(skuTitle))
    if 1 <= var.nn:
      var.nn -= 1
      if var.nn != 0:
        var.n1 = var.skuTitle[var.nn-1]
      else:
        var.n1 = ""
      var.n2 = var.skuTitle[var.nn]
      var.n3 = var.skuTitle[var.nn+1]
    else:
      var.n1 = ""
      var.n2 = var.skuTitle[var.nn]
      var.n3 = var.skuTitle[var.nn+1]
    self.frames2a()
    self.frames2b()
    mainFunc().changetxtdetail2()

  def next(self):
    print("dest frames")
    self.leftframe.destroy()
    self.rightframe.destroy()
    print("frames destroyed")
    if var.skulength-1 == var.nn:
      var.n1 = var.skuTitle[var.nn-1]
      var.n2 = var.skuTitle[var.nn]
      var.n3 = ""
    else:
      var.nn += 1
      var.n1 = var.skuTitle[var.nn-1]
      var.n2 = var.skuTitle[var.nn]
      if var.skulength == var.nn+1:
        var.n3 = ""
      else:
        var.n3 = var.skuTitle[var.nn+1]
    self.frames2a()
    self.frames2b()
    mainFunc().changetxtdetail2()

  def opt1(self):
    var.results[var.nn] = ["FALSE", var.essVal[var.nn][0]]
    #print(var.results[var.nn])
    self.prevID = var.essVal[var.nn][0]
    self.next()
  def opt2(self):
    var.results[var.nn] =["FALSE", var.essVal[var.nn][1]]
    #print(var.results[var.nn])
    self.prevID = var.essVal[var.nn][1]
    self.next()
  def opt3(self):
    var.results[var.nn] =["FALSE", var.essVal[var.nn][2]]
    #print(var.results[var.nn])
    self.prevID = var.essVal[var.nn][2]
    self.next()

  def newTrue(self):
    var.results[var.nn] =["TRUE", var.skuID[var.nn]]
    print(var.results[var.nn])
    self.prevID = var.skuID[var.nn]
    self.next()
  def dupPrev(self):
    var.results[var.nn] =["FALSE", self.prevID]
    print(var.results[var.nn])
    self.next()
  def blanklabel(self):
    var.results[var.nn] = ["FALSE", "###!"]
    self.prevID = "###!"
    print(var.results[var.nn])
    self.next()

  def gonext(self):
    n = 1
    if var.results[var.nn] == []:
      print("finding next labelled")
      for gn in var.results[var.nn:]:
        if gn != []:
          var.nn = n + var.nn
          self.previous()
          break
        n += 1
    else:
      print("finding next empty")
      for gn in var.results[var.nn:]:
        if gn == []:
          var.nn = n + var.nn
          self.previous()
          break
        n += 1

  def goprev(self):
    n = var.nn
    if var.results[var.nn] == []:
      print("finding prev labelled")
      for gn in reversed(var.results[:var.nn]):
        if gn != []:
          var.nn = n
          self.previous()
          break
        n -= 1
    else:
      print("finding prev empty")
      for gn in reversed(var.results[:var.nn]):
        if gn == []:
          var.nn = n
          self.previous()
          break
        n -= 1




  def save(self):
    print("saving pickle data")
    dbfile = open('labelled', 'wb')
    db = var.results
    print(var.results[:10])
    pickle.dump(db, dbfile)
    dbfile.close()

    print("reading pickle label 10 lines")
    dbfile = open('labelled', 'rb')
    print(pickle.load(dbfile)[:10])
    dbfile.close()
    print("saving pickle data completed")

    print("Saving to Text file too")
    self.savetxt()


  def savetxt(self):
    f = open("labelledA.txt", "w")
    try:
      resl = re.sub(r"^.*?\[\'", "", str(var.results))
      resl = re.sub(r"\'(?:.(?!\'))+$", "", resl)
      resl = resl.replace("], [","\n")
      resl = resl.replace("', '","\t")
      resl = re.sub(r"^\'", "", resl, flags=re.MULTILINE)
      resl = re.sub(r"\'$", "", resl, flags=re.MULTILINE)
      f.write(resl)
      f.close()
      print("saved, next save will overwrite")
    except:
      print("Error while reformatting to save the txt file [labelA]")
    f = open("labelledB.txt", "w")
    try:
      f.write(str(var.results))
      f.close()
      print("saved, next save will overwrite")
    except:
      print("Error while reformatting to save the txt file [labelB]")

  def reload(self):
    mainFunc().loaddata()
    var.nn = 1
    self.previous()

  def lastwork(self):
    for idx, lw in enumerate(reversed(var.results)):
      if lw != []:
        var.nn = var.skulength-idx
        print(lw, var.nn)
        self.previous()
        break


root = tk.Tk()
root.title("Usman")
mainFunc().loaddata()
tkApp(root)
root.mainloop()
